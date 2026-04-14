#!/usr/bin/env python3
"""
Build viewer data from papers_record.xlsx.

Usage:
  /home/wsg/.hermes/hermes-agent/venv/bin/python build_data.py
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = BASE_DIR / "papers_record.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parent / "papers_data.json"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def load_rows() -> list[dict]:
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["Papers"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header_row]
    index = {name: i for i, name in enumerate(headers)}

    required = [
        "arxiv_id",
        "title",
        "authors",
        "affiliations",
        "published_date",
        "categories",
        "abstract",
        "summary_cn",
        "pdf_filename",
        "crawled_date",
        "notes",
    ]
    missing = [c for c in required if c not in index]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    def quality_key(p: dict) -> tuple:
        return (
            1 if p.get("summary_cn") else 0,
            1 if p.get("affiliations") else 0,
            len(p.get("summary_cn", "")),
            len(p.get("affiliations", "")),
            len(p.get("abstract", "")),
            p.get("crawled_date", ""),
            p.get("published_date", ""),
        )

    rows_by_id: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        paper = {col: normalize_text(row[index[col]]) for col in required}
        if not paper["arxiv_id"]:
            continue
        paper["pdf_url"] = f"https://arxiv.org/pdf/{paper['arxiv_id']}"
        arxiv_id = paper["arxiv_id"]
        old = rows_by_id.get(arxiv_id)
        if old is None or quality_key(paper) > quality_key(old):
            rows_by_id[arxiv_id] = paper

    rows = list(rows_by_id.values())

    rows.sort(key=lambda x: (x["crawled_date"], x["published_date"], x["arxiv_id"]), reverse=True)
    return rows


def main() -> None:
    papers = load_rows()

    crawled_dates = sorted({p["crawled_date"] for p in papers if p["crawled_date"]})
    published_dates = sorted({p["published_date"] for p in papers if p["published_date"]})

    payload = {
        "count": len(papers),
        "crawled_date_min": crawled_dates[0] if crawled_dates else "",
        "crawled_date_max": crawled_dates[-1] if crawled_dates else "",
        "published_date_min": published_dates[0] if published_dates else "",
        "published_date_max": published_dates[-1] if published_dates else "",
        "papers": papers,
    }

    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] Wrote {len(papers)} papers to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
